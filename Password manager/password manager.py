from tkinter import *
from tkinter import ttk
import openpyxl
#HEHEHEHE

#Okno
screen = Tk()
screen.title("Password manager")
screen.minsize(800, 700)
screen.resizable(True, True)

#Funkce


def adding():
    list_box.insert("", END, values=(add_input_name.get(), add_input_password.get(), add_input_note.get()))

def deleting_item():
    list_box.delete(list_box.focus())

def saving_items():
    # otevřít soubor Excel nebo vytvořit nový
    try:
        workbook = openpyxl.load_workbook("items.xlsx")
    except:
        workbook = openpyxl.Workbook()

    # vybrat nebo vytvořit list
    try:
        sheet = workbook["Items"]
    except:
        sheet = workbook.create_sheet("Items")




    items = list_box.get_children()

    for item in items:
        name, password, note = list_box.item(item)['values']
        sheet.append([name, password, note])

    workbook.save("items.xlsx")

def loading_items():
    # otevřít soubor Excel
    try:
        workbook = openpyxl.load_workbook("items.xlsx")
    except:
        return

    # vybrat list
    try:
        sheet = workbook["Items"]
    except:
        return

    # načíst položky z listu
    for row in sheet.iter_rows(min_row=2, values_only=True):
        list_box.insert("", END, values=row)


def quitting_app():
    screen.quit()


#Framy
adding_frame = Frame(screen)
adding_frame.pack()

list_frame = Frame(screen)
list_frame.pack()

button_frame = Frame(screen)
button_frame.pack()


#Přidání oblasti

name_label = Label(adding_frame, text="Název")
name_label.grid(row=0, column=0)

password_label = Label(adding_frame, text="Heslo")
password_label.grid(row=0, column=1)

note_label = Label(adding_frame, text="Poznámka")
note_label.grid(row=0, column=2)



add_input_name = Entry(adding_frame, width=30)
add_input_name.grid(row=1, column=0, pady=5, padx=5)

add_input_password = Entry(adding_frame, width=30)
add_input_password.grid(row=1, column=1, pady=5, padx=5)

add_input_note = Entry(adding_frame, width=30)
add_input_note.grid(row=1, column=2, pady=5, padx=5)

name_input = Button(adding_frame, text="Přidat jméno", command=adding)
name_input.grid(row=1, column=3, pady=5, padx=5)


#List oblast
list_box = ttk.Treeview(list_frame, height=15)
list_box.grid(row=1, column=0)


list_box["columns"] = ("1", "2", "3")

list_box.column("#0", width=0, stretch=FALSE)

list_box.column("1", width=250, stretch=FALSE)
list_box.heading("1", text="Název")

list_box.column("2", width=250)
list_box.heading("2", text="Heslo")

list_box.column("3", width=250)
list_box.heading("3", text="Poznámka")


#Tlačítka
removeItem_button = Button(button_frame, text="Odstranit položku", command=deleting_item)
removeItem_button.grid(row=2, column=0, pady=5, padx=5)

deleteList_button = Button(button_frame, text="Odstranit seznam")
deleteList_button.grid(row=2, column=1, pady=5, padx=5)

saveItem_button = Button(button_frame, text="Uložit položku", command=saving_items)
saveItem_button.grid(row=2, column=2, pady=5, padx=5)

quit_button = Button(button_frame, text="Ukončit aplikace", command=quitting_app)
quit_button.grid(row=2, column=3, pady=5, padx=5)

loading_items()

#Hlavní cyklus
screen.mainloop()